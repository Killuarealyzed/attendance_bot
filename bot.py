import asyncio
import sqlite3
import re
import os
from datetime import datetime, timedelta
from dotenv import load_dotenv

# ===== EXCEL ИНТЕГРАЦИЯ =====
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ===== ПЛАНИРОВЩИК ЗАДАЧ =====
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from zoneinfo import ZoneInfo

from aiogram import Bot, Dispatcher, Router
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, FSInputFile
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.exceptions import TelegramForbiddenError, TelegramRetryAfter, TelegramAPIError

# ===== КОНСТАНТЫ =====
EXCEL_FILE = "attendance_journal.xlsx"
load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_CHAT_ID_RAW = os.getenv("ADMIN_CHAT_ID")

if not BOT_TOKEN or BOT_TOKEN.strip() == "":
    raise ValueError("❌ ОШИБКА: Не найден BOT_TOKEN в файле .env!")
if not ADMIN_CHAT_ID_RAW or ADMIN_CHAT_ID_RAW.strip() == "":
    raise ValueError("❌ ОШИБКА: Не найден ADMIN_CHAT_ID в файле .env!")

try:
    ADMIN_CHAT_ID = int(ADMIN_CHAT_ID_RAW.strip())
except ValueError:
    raise ValueError(f"❌ ОШИБКА: ADMIN_CHAT_ID должен быть числом, получено: '{ADMIN_CHAT_ID_RAW}'")

# ===== ИНИЦИАЛИЗАЦИЯ БОТА =====
bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
router = Router()

# ===== FSM СОСТОЯНИЯ =====
class AttendanceForm(StatesGroup):
    waiting_for_name = State()
    waiting_for_attendance = State()
    waiting_for_date = State()
    waiting_for_reason = State()
    waiting_for_start_date = State()
    waiting_for_end_date = State()
    waiting_for_absence_reason = State()

# ===== ФУНКЦИИ ДЛЯ РАБОТЫ С ДАТАМИ =====
def get_weekdays(start_date: datetime, days_ahead: int = 30) -> list:
    """Генерирует список учебных дней (пн-сб)"""
    weekdays = []
    current_date = start_date
    for _ in range(days_ahead):
        if current_date.weekday() < 6:  # пн-сб
            weekdays.append(current_date.strftime("%d.%m.%Y"))
        current_date += timedelta(days=1)
    return weekdays

def parse_date(date_str: str) -> datetime:
    """Преобразует строку ДД.ММ или ДД.ММ.ГГГГ в datetime"""
    parts = date_str.split('.')
    if len(parts) == 2:
        day, month = int(parts[0]), int(parts[1])
        year = datetime.now().year
        if datetime(year, month, day) < datetime.now():
            year += 1
    else:
        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
    return datetime(year, month, day)

def get_date_range(start_date: datetime, end_date: datetime) -> list:
    """Возвращает все учебные дни в диапазоне [start_date, end_date]"""
    dates = []
    current = start_date
    while current <= end_date:
        if current.weekday() < 6:  # пн-сб
            dates.append(current.strftime("%d.%m.%Y"))
        current += timedelta(days=1)
    return dates

def ensure_dates_in_excel(ws, start_date: datetime = None, days_ahead: int = 30):
    """Гарантирует наличие всех необходимых дат в Excel"""
    if start_date is None:
        start_date = datetime.now()
    
    existing_dates = set()
    for col in range(4, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            existing_dates.add(str(cell_value))
    
    needed_dates = get_weekdays(start_date, days_ahead)
    new_dates_added = 0
    
    for date_str in needed_dates:
        if date_str not in existing_dates:
            insert_col = 4
            for col in range(4, ws.max_column + 1):
                existing_date = ws.cell(row=1, column=col).value
                if existing_date:
                    try:
                        existing_dt = datetime.strptime(str(existing_date), "%d.%m.%Y")
                        needed_dt = datetime.strptime(date_str, "%d.%m.%Y")
                        if needed_dt < existing_dt:
                            insert_col = col
                            break
                    except:
                        pass
                insert_col = col + 1
            
            ws.insert_cols(insert_col)
            ws.cell(row=1, column=insert_col, value=date_str)
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
            ws.cell(row=1, column=insert_col).font = header_font
            ws.cell(row=1, column=insert_col).fill = header_fill
            ws.cell(row=1, column=insert_col).alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(insert_col)].width = 15
            new_dates_added += 1
    
    if new_dates_added > 0:
        print(f"✅ Добавлено {new_dates_added} новых учебных дат (пн-сб) в журнал")
    return new_dates_added

# ===== ИНИЦИАЛИЗАЦИЯ БАЗЫ ДАННЫХ =====
def init_db():
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                username TEXT,
                last_active TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS absences (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                reason TEXT,
                reported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(user_id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS absence_periods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                reason TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(user_id)
            )
        ''')
        
        cursor.execute("PRAGMA table_info(users)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'username' not in columns:
            cursor.execute("ALTER TABLE users ADD COLUMN username TEXT")
        
        conn.commit()
        conn.close()
        print("✅ База данных инициализирована")
    except Exception as e:
        print(f"❌ Ошибка инициализации БД: {e}")
        raise

# ===== EXCEL ФУНКЦИИ =====
def init_excel():
    """Создаёт Excel-файл с базовой структурой"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал посещаемости"
    
    ws['A1'] = "ID"
    ws['B1'] = "Имя"
    ws['C1'] = "Юзернейм"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col in ['A', 'B', 'C']:
        ws[f"{col}1"].font = header_font
        ws[f"{col}1"].fill = header_fill
        ws[f"{col}1"].alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    
    ensure_dates_in_excel(ws, datetime.now(), 30)
    wb.save(EXCEL_FILE)
    print(f"✅ Создан Excel-файл: {EXCEL_FILE}")

def ensure_user_in_excel(user_id: int, name: str, username: str = None):
    """Гарантирует, что пользователь есть в Excel."""
    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()
            
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        user_exists = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                user_exists = True
                ws.cell(row=row, column=2, value=name)
                ws.cell(row=row, column=3, value=f"@{username}" if username else "")
                break
        
        if not user_exists:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=user_id)
            ws.cell(row=new_row, column=2, value=name)
            ws.cell(row=new_row, column=3, value=f"@{username}" if username else "")
            print(f"✅ Добавлен новый пользователь в Excel: {name} (ID: {user_id})")
        
        wb.save(EXCEL_FILE)
        return True
        
    except Exception as e:
        print(f"❌ Ошибка проверки пользователя в Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def update_attendance_in_excel(user_id: int, date_str: str, status: str, reason: str = None):
    """Обновляет посещаемость в Excel."""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT name, username FROM users WHERE user_id = ?", (user_id,))
        user_data = cursor.fetchone()
        conn.close()
        
        if not user_data:
            print(f"⚠️ Пользователь ID {user_id} не найден в БД")
            return
        
        name, username = user_data
        ensure_user_in_excel(user_id, name, username)
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ensure_dates_in_excel(ws, datetime.now(), 30)
        
        date_col = None
        for col in range(4, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if str(cell_value) == date_str:
                date_col = col
                break
        
        if date_col is None:
            print(f"❌ Дата {date_str} не найдена в Excel")
            return
        
        user_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                user_row = row
                break
        
        if user_row is None:
            print(f"❌ Не удалось найти пользователя ID {user_id} в Excel")
            return
        
        status_text = status
        if reason and status == "❌":
            status_text += f"\n({reason})"
        
        ws.cell(row=user_row, column=date_col, value=status_text)
        ws.cell(row=user_row, column=date_col).alignment = Alignment(wrap_text=True, horizontal="center")
        
        if status == "✅":
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=user_row, column=date_col).fill = fill
        wb.save(EXCEL_FILE)
        print(f"✅ Обновлена посещаемость: ID {user_id}, дата {date_str}, статус {status}")
        
    except Exception as e:
        print(f"❌ Ошибка обновления Excel: {e}")
        import traceback
        traceback.print_exc()

# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====
def get_main_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📝 Отметиться")],
            [KeyboardButton(text="📆 Отсутствую с... по...")]
        ],
        resize_keyboard=True,
        one_time_keyboard=False
    )

def get_cancel_kb():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="🚫 Отмена")]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

def validate_and_normalize_date(date_str: str) -> tuple[bool, str]:
    date_str = date_str.strip()
    if not re.match(r'^\d{1,2}\.\d{1,2}(\.\d{4})?$', date_str):
        return False, "Неверный формат. Используй ДД.ММ или ДД.ММ.ГГГГ (например, 15.02 или 15.02.2026)"
    
    try:
        dt = parse_date(date_str)
        return True, dt.strftime("%d.%m.%Y")
    except:
        return False, "Некорректная дата"

def is_user_absent_today(user_id: int, today: str) -> bool:
    """Проверяет, находится ли пользователь в периоде отсутствия сегодня"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id FROM absence_periods 
            WHERE user_id = ? 
            AND ? BETWEEN start_date AND end_date
        """, (user_id, today))
        result = cursor.fetchone()
        conn.close()
        return result is not None
    except:
        return False

# ===== ХЕНДЛЕРЫ КОМАНД (ОБЯЗАТЕЛЬНО В НАЧАЛЕ!) =====
@router.message(Command("help"))
async def cmd_help(message: Message):
    help_text = (
        "ℹ️ Команды:\n"
        "/start — начать диалог\n"
        "/history — история отсутствий\n"
        "/absence — активные периоды отсутствия\n"
        "/clear_absence — удалить периоды\n"
        "/journal — получить Excel-журнал (админ)\n\n"
        "📅 Учебные дни: понедельник-суббота"
    )
    await message.answer(help_text)

@router.message(Command("history"))
async def cmd_history(message: Message):
    user_id = message.from_user.id
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT date, reason FROM absences WHERE user_id = ? ORDER BY rowid DESC LIMIT 10", (user_id,))
        absences = cursor.fetchall()
        conn.close()
    except Exception as e:
        await message.answer(f"❌ Ошибка БД: {e}")
        return

    if not absences:
        await message.answer("📭 Нет записанных отсутствий.")
        return

    text = "📊 История отсутствий:\n\n"
    for date, reason in absences:
        reason_str = f" — {reason}" if reason else ""
        text += f"• {date}{reason_str}\n"
    await message.answer(text)

@router.message(Command("absence"))
async def cmd_absence(message: Message):
    """Показывает активные периоды отсутствия пользователя"""
    user_id = message.from_user.id
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT start_date, end_date, reason 
            FROM absence_periods 
            WHERE user_id = ? AND end_date >= ?
            ORDER BY start_date
        """, (user_id, datetime.now().strftime("%d.%m.%Y")))
        periods = cursor.fetchall()
        conn.close()
        
        if not periods:
            await message.answer("📭 У вас нет активных периодов отсутствия.")
            return
        
        text = "📅 Активные периоды отсутствия:\n\n"
        for start_date, end_date, reason in periods:
            text += f"📆 С {start_date} по {end_date}\n📝 {reason}\n\n"
        
        text += "💡 Чтобы удалить период, отправьте /clear_absence"
        await message.answer(text)
        
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")

@router.message(Command("clear_absence"))
async def cmd_clear_absence(message: Message):
    """Удаляет все активные периоды отсутствия пользователя"""
    user_id = message.from_user.id
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM absence_periods 
            WHERE user_id = ? AND end_date >= ?
        """, (user_id, datetime.now().strftime("%d.%m.%Y")))
        deleted = cursor.rowcount
        conn.commit()
        conn.close()
        
        if deleted > 0:
            await message.answer(f"✅ Удалено {deleted} активных периодов отсутствия.")
        else:
            await message.answer("📭 Нет активных периодов для удаления.")
            
    except Exception as e:
        await message.answer(f"❌ Ошибка удаления: {e}")

@router.message(Command("journal"))
async def cmd_journal(message: Message):
    if message.from_user.id != ADMIN_CHAT_ID:
        await message.answer("❌ Эта команда только для админа!")
        return
    
    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ensure_dates_in_excel(ws, datetime.now(), 30)
            wb.save(EXCEL_FILE)
        
        document = FSInputFile(EXCEL_FILE, filename="Журнал_посещаемости.xlsx")
        await message.answer_document(document, caption="📊 Актуальный журнал посещаемости")
    except Exception as e:
        await message.answer(f"❌ Ошибка отправки файла: {e}")
        import traceback
        traceback.print_exc()

# ===== ХЕНДЛЕР /start (только при команде) =====
@router.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    user_id = message.from_user.id
    username = message.from_user.username
    
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT name, username FROM users WHERE user_id = ?", (user_id,))
        user = cursor.fetchone()
        
        if user and username != user[1]:
            cursor.execute("UPDATE users SET username = ? WHERE user_id = ?", (username, user_id))
            conn.commit()
        conn.close()
    except Exception as e:
        await message.answer(f"❌ Ошибка базы данных: {e}")
        return

    if user:
        await message.answer(
            f"👋 Привет, {user[0]}!\n\nВыбери действие:",
            reply_markup=get_main_kb()
        )
        await state.clear()  # ← КРИТИЧЕСКИ ВАЖНО!
    else:
        await message.answer("👋 Представься (ФИО или имя):", reply_markup=ReplyKeyboardRemove())
        await state.set_data({"username": username})
        await state.set_state(AttendanceForm.waiting_for_name)

# ===== ГЛОБАЛЬНЫЙ ХЕНДЛЕР ДЛЯ КНОПОК (только если состояние пустое) =====
@router.message(
    lambda message: message.text in ["📝 Отметиться", "📆 Отсутствую с... по..."],
    StateFilter(None)  # ← Только если состояние не установлено
)
async def handle_buttons(message: Message, state: FSMContext):
    user_id = message.from_user.id
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM users WHERE user_id = ?", (user_id,))
        user = cursor.fetchone()
        conn.close()
        
        if not user:
            await message.answer("Сначала представьтесь! Нажмите /start")
            return
            
        if message.text == "📝 Отметиться":
            await message.answer(
                "Выбери свой статус на сегодня:",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[
                        [KeyboardButton(text="✅ Буду"), KeyboardButton(text="❌ Не буду")]
                    ],
                    resize_keyboard=True,
                    one_time_keyboard=True
                )
            )
            await state.set_state(AttendanceForm.waiting_for_attendance)
            
        elif message.text == "📆 Отсутствую с... по...":
            await message.answer(
                "📅 Укажи дату начала отсутствия (ДД.ММ.ГГГГ):",
                reply_markup=get_cancel_kb()
            )
            await state.set_state(AttendanceForm.waiting_for_start_date)
            
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")

# ===== ХЕНДЛЕРЫ FSM (обработка состояний) =====
@router.message(AttendanceForm.waiting_for_name)
async def process_name(message: Message, state: FSMContext):
    name = message.text.strip()
    if len(name) < 2:
        await message.answer("❌ Слишком короткое имя. Попробуй ещё:")
        return
    
    user_id = message.from_user.id
    username = message.from_user.username or (await state.get_data()).get("username")
    
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("INSERT OR REPLACE INTO users (user_id, name, username) VALUES (?, ?, ?)", (user_id, name, username))
        conn.commit()
        conn.close()
        ensure_user_in_excel(user_id, name, username)
    except Exception as e:
        await message.answer(f"❌ Ошибка сохранения: {e}")
        return
    
    await message.answer(
        f"✅ Привет, {name}!\n\nВыбери действие:",
        reply_markup=get_main_kb()
    )
    await state.clear()  # ← Сбрасываем состояние после регистрации

@router.message(AttendanceForm.waiting_for_attendance)
async def process_attendance(message: Message, state: FSMContext):
    today = datetime.now().strftime("%d.%m.%Y")
    
    if message.text == "✅ Буду":
        user_id = message.from_user.id
        update_attendance_in_excel(user_id, today, "✅")
        await message.answer("👍 Отлично! Хороших пар! 📚", reply_markup=get_main_kb())
        await state.clear()
        return
        
    elif message.text == "❌ Не буду":
        await message.answer(
            "📅 Укажи дату отсутствия (ДД.ММ.ГГГГ):",
            reply_markup=get_cancel_kb()
        )
        await state.set_state(AttendanceForm.waiting_for_date)
        return
        
    elif message.text == "🚫 Отмена":
        await message.answer("↩️ Отменено.", reply_markup=get_main_kb())
        await state.clear()
        return
        
    await message.answer("❓ Используй кнопки 👇", reply_markup=get_main_kb())

@router.message(AttendanceForm.waiting_for_date)
async def process_date(message: Message, state: FSMContext):
    if message.text == "🚫 Отмена":
        await message.answer("↩️ Отменено.", reply_markup=get_main_kb())
        await state.clear()
        return
    
    is_valid, result = validate_and_normalize_date(message.text)
    if not is_valid:
        await message.answer(f"❌ {result}\nПопробуй ещё:")
        return
    
    await state.update_data(date=result)
    await message.answer("✏️ Причина отсутствия? Напиши «-» если нет:", reply_markup=get_cancel_kb())
    await state.set_state(AttendanceForm.waiting_for_reason)

@router.message(AttendanceForm.waiting_for_reason)
async def process_reason(message: Message, state: FSMContext):
    if message.text == "🚫 Отмена":
        await message.answer("↩️ Отменено.", reply_markup=get_main_kb())
        await state.clear()
        return
    
    reason = None if message.text.strip() in ["-", ""] else message.text.strip()
    user_id = message.from_user.id
    data = await state.get_data()
    date = data['date']
    
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT name, username FROM users WHERE user_id = ?", (user_id,))
        user_row = cursor.fetchone()
        if not user_row:
            conn.close()
            await message.answer("❌ Ошибка: пользователь не найден в базе.")
            await state.clear()
            return
        
        user_name, user_username = user_row
        cursor.execute("INSERT INTO absences (user_id, date, reason) VALUES (?, ?, ?)", (user_id, date, reason))
        conn.commit()
        conn.close()
    except Exception as e:
        await message.answer(f"❌ Ошибка сохранения: {e}")
        await state.clear()
        return
    
    update_attendance_in_excel(user_id, date, "❌", reason)
    
    username_display = f" (@{user_username})" if user_username else ""
    reason_text = f"\n📝 Причина: {reason}" if reason else ""
    await bot.send_message(
        ADMIN_CHAT_ID,
        f"⚠️ Отсутствие\n👤 {user_name}{username_display} (ID: {user_id})\n📅 {date}{reason_text}"
    )
    
    await message.answer(f"✅ Записал отсутствие на {date}.", reply_markup=get_main_kb())
    await state.clear()

@router.message(AttendanceForm.waiting_for_start_date)
async def process_start_date(message: Message, state: FSMContext):
    if message.text == "🚫 Отмена":
        await message.answer("↩️ Отменено.", reply_markup=get_main_kb())
        await state.clear()
        return
    
    is_valid, result = validate_and_normalize_date(message.text)
    if not is_valid:
        await message.answer(f"❌ {result}\nПопробуй ещё:")
        return
    
    await state.update_data(start_date=result)
    await message.answer(
        "📅 Укажи дату окончания отсутствия (ДД.ММ или ДД.ММ.ГГГГ):",
        reply_markup=get_cancel_kb()
    )
    await state.set_state(AttendanceForm.waiting_for_end_date)

@router.message(AttendanceForm.waiting_for_end_date)
async def process_end_date(message: Message, state: FSMContext):
    if message.text == "🚫 Отмена":
        await message.answer("↩️ Отменено.", reply_markup=get_main_kb())
        await state.clear()
        return
    
    is_valid, result = validate_and_normalize_date(message.text)
    if not is_valid:
        await message.answer(f"❌ {result}\nПопробуй ещё:")
        return
    
    data = await state.get_data()
    start_date = data['start_date']
    
    try:
        start_dt = datetime.strptime(start_date, "%d.%m.%Y")
        end_dt = datetime.strptime(result, "%d.%m.%Y")
        if end_dt < start_dt:
            await message.answer("❌ Дата окончания не может быть раньше даты начала!\nУкажи корректную дату окончания:")
            return
    except:
        await message.answer("❌ Ошибка при сравнении дат. Попробуй ещё:")
        return
    
    await state.update_data(end_date=result)
    await message.answer("✏️ Укажи причину отсутствия (болезнь, отпуск и т.д.):", reply_markup=get_cancel_kb())
    await state.set_state(AttendanceForm.waiting_for_absence_reason)

@router.message(AttendanceForm.waiting_for_absence_reason)
async def process_absence_reason(message: Message, state: FSMContext):
    if message.text == "🚫 Отмена":
        await message.answer("↩️ Отменено.", reply_markup=get_main_kb())
        await state.clear()
        return
    
    reason = message.text.strip()
    user_id = message.from_user.id
    data = await state.get_data()
    start_date = data['start_date']
    end_date = data['end_date']
    
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT name, username FROM users WHERE user_id = ?", (user_id,))
        user_row = cursor.fetchone()
        if not user_row:
            conn.close()
            await message.answer("❌ Ошибка: пользователь не найден в базе.")
            await state.clear()
            return
        
        user_name, user_username = user_row
        cursor.execute(
            "INSERT INTO absence_periods (user_id, start_date, end_date, reason) VALUES (?, ?, ?, ?)",
            (user_id, start_date, end_date, reason)
        )
        conn.commit()
        conn.close()
        
        date_range = get_date_range(
            datetime.strptime(start_date, "%d.%m.%Y"),
            datetime.strptime(end_date, "%d.%m.%Y")
        )
        
        for date_str in date_range:
            update_attendance_in_excel(user_id, date_str, "❌", reason)
        
        username_display = f" (@{user_username})" if user_username else ""
        admin_message = (
            f"📅 ПЕРИОД ОТСУТСТВИЯ\n\n"
            f"👤 {user_name}{username_display} (ID: {user_id})\n"
            f"📆 С {start_date} по {end_date}\n"
            f"📝 Причина: {reason}"
        )
        await bot.send_message(ADMIN_CHAT_ID, admin_message)
        
        await message.answer(
            f"✅ Записал период отсутствия:\n"
            f"📆 С {start_date} по {end_date}\n"
            f"📝 Причина: {reason}\n\n"
            f"Бот не будет беспокоить вас в эти дни!",
            reply_markup=get_main_kb()
        )
        await state.clear()
        
    except Exception as e:
        await message.answer(f"❌ Ошибка сохранения периода: {e}")
        await state.clear()

# ===== ФУНКЦИЯ ЕЖЕДНЕВНОГО НАПОМИНАНИЯ В 20:00 =====
async def send_daily_reminder(bot: Bot):
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, name, username FROM users")
        users = cursor.fetchall()
        conn.close()
        
        if not users:
            print("📭 Нет зарегистрированных пользователей")
            return
        
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%d.%m.%Y")
        success_count = 0
        
        for user_id, name, username in users:
            if is_user_absent_today(user_id, tomorrow):
                print(f"⏭️ Пропускаем пользователя {name} (ID: {user_id}) — в отпуске завтра")
                continue
                
            try:
                username_display = f" (@{username})" if username else ""
                message_text = (
                    f"🌙 Вечернее напоминание\n\n"
                    f"{name}{username_display}, будешь завтра на парах?\n\n"
                    f"📅 Завтра: {tomorrow}"
                )
                
                await bot.send_message(user_id, message_text, reply_markup=get_main_kb())
                success_count += 1
                await asyncio.sleep(0.05)
                
            except (TelegramForbiddenError, TelegramAPIError):
                continue
        
        print(f"✅ Напоминание отправлено {success_count} пользователям")
        
    except Exception as e:
        print(f"❌ Ошибка напоминания: {e}")
        import traceback
        traceback.print_exc()

# ===== ГЛАВНАЯ ФУНКЦИЯ =====
async def main():
    print(f"🔧 Админский ID: {ADMIN_CHAT_ID}")
    print(f"🤖 Запуск бота...")
    
    init_db()
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ensure_dates_in_excel(ws, datetime.now(), 30)
        wb.save(EXCEL_FILE)
    
    dp.include_router(router)
    await bot.set_my_commands([
        {"command": "start", "description": "Начать диалог"},
        {"command": "history", "description": "История отсутствий"},
        {"command": "absence", "description": "Периоды отсутствия"},
        {"command": "clear_absence", "description": "Удалить периоды"},
        {"command": "help", "description": "Помощь"},
        {"command": "journal", "description": "Получить журнал (админ)"},
    ])
    
    scheduler = AsyncIOScheduler(timezone=ZoneInfo("Europe/Moscow"))
    scheduler.add_job(
        send_daily_reminder,
        CronTrigger(hour=20, minute=0, timezone=ZoneInfo("Europe/Moscow")),
        args=[bot],
        id="evening_reminder",
        replace_existing=True,
        misfire_grace_time=1800
    )
    scheduler.start()
    print("⏰ Планировщик запущен: напоминание в 20:00 по МСК")
    print("📅 Учтены учебные дни: понедельник-суббота")
    print(f"📊 Excel-журнал: {os.path.abspath(EXCEL_FILE)}")
    
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n👋 Бот остановлен.")
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()